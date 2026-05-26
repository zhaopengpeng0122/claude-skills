#!/usr/bin/env python3
"""业务线测试任务总清单 xlsx → HTML 生成器"""
import sys, os, json, re
import openpyxl

def read_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # Sheet 1: 任务表
    ws = wb['产品线全量测试任务表']
    # 找实际使用列数
    max_col = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None:
                max_col = max(max_col, cell.column)
    max_col = min(max_col, 22)  # 最多取到 V 列

    tasks = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=max_col, values_only=True):
        vals = list(row) + [''] * (22 - len(list(row)))
        if vals[0] is None:
            continue
        tasks.append({
            'layer': vals[0] or '',
            'module': vals[1] or '',
            'testType': vals[2] or '',
            'testTask': vals[3] or '',
            'taskType': vals[4] or '',
            'taskId': vals[5] or '',
            'taskName': vals[6] or '',
            'executor': vals[7] or '',
            'sceneDesc': vals[8] or '',
            'testSetSummary': vals[9] or '',
            'testEnv': vals[10] or '',
            'testTool': vals[11] or '',
            'metricFormula': vals[12] or '',
            'contractTarget': vals[13] or '',
            'competeTarget': vals[14] or '',
            'kaTarget': vals[15] or '',
            'score5': vals[16] or '',
            'score4': vals[17] or '',
            'score3': vals[18] or '',
            'score2': vals[19] or '',
            'score1': vals[20] or '',
            'score0': vals[21] or '',
            'actualResult': '',
            'judgeMode': '正常评分（Normal）',
        })

    # Weights
    weights = {'taskType': [], 'testType': [], 'module': [], 'layer': []}

    ws2 = wb['任务权重定义']
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
        if row[0]: weights['taskType'].append({'name': str(row[0]), 'weight': row[1]})
        if len(row) > 4 and row[4]: weights['testType'].append({'name': str(row[4]), 'weight': row[5]})

    ws3 = wb['模块权重定义']
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, values_only=True):
        if row[0]: weights['module'].append({'name': str(row[0]), 'weight': row[1]})

    ws4 = wb['业务层权重定义']
    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, values_only=True):
        if row[0]: weights['layer'].append({'name': str(row[0]), 'weight': row[1]})

    return {'tasks': tasks, 'weights': weights}


def extract_project_name(path):
    basename = os.path.splitext(os.path.basename(path))[0]
    # 去掉 "业务线测试任务总清单-" 前缀 或 "-业务线测试任务总清单" 后缀
    name = re.sub(r'^业务线测试任务总清单-', '', basename)
    name = re.sub(r'-业务线测试任务总清单$', '', name)
    return name


def generate_html(data, project_name):
    json_data = json.dumps(data, ensure_ascii=False, indent=None)

    return f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{project_name} - 业务线测试任务总清单</title>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, "Microsoft YaHei", sans-serif; background: #f0f2f5; color: #1d2129; }}
.header {{ background: linear-gradient(135deg, #1a73e8, #0d47a1); color: #fff; padding: 20px 32px; display: flex; align-items: center; justify-content: space-between; }}
.header h1 {{ font-size: 22px; font-weight: 600; }}
.header .stats {{ font-size: 14px; opacity: .9; }}

.toolbar {{ background: #fff; padding: 16px 24px; display: flex; gap: 12px; flex-wrap: wrap; align-items: center; border-bottom: 1px solid #e5e6eb; }}
.toolbar label {{ font-size: 13px; color: #4e5969; margin-right: 4px; }}
.toolbar select, .toolbar input[type="text"] {{ height: 32px; border: 1px solid #c9cdd4; border-radius: 4px; padding: 0 8px; font-size: 13px; outline: none; }}
.toolbar select:focus, .toolbar input:focus {{ border-color: #1a73e8; box-shadow: 0 0 0 2px rgba(26,115,232,.15); }}
.toolbar input[type="text"] {{ width: 160px; }}
.btn {{ height: 32px; padding: 0 14px; border: none; border-radius: 4px; font-size: 13px; cursor: pointer; display: inline-flex; align-items: center; gap: 4px; }}
.btn-primary {{ background: #1a73e8; color: #fff; }}
.btn-primary:hover {{ background: #1557b0; }}
.btn-success {{ background: #00b42a; color: #fff; }}
.btn-success:hover {{ background: #009a25; }}
.btn-danger {{ background: #f53f3f; color: #fff; }}
.btn-danger:hover {{ background: #d91a1a; }}
.btn-outline {{ background: #fff; color: #4e5969; border: 1px solid #c9cdd4; }}
.btn-outline:hover {{ border-color: #1a73e8; color: #1a73e8; }}
.spacer {{ flex: 1; }}

.table-wrap {{ padding: 16px 24px; }}
.table-info {{ font-size: 13px; color: #86909c; margin-bottom: 8px; }}
table {{ width: 100%; border-collapse: collapse; background: #fff; border-radius: 6px; overflow: hidden; box-shadow: 0 1px 4px rgba(0,0,0,.08); }}
thead th {{ background: #f2f3f5; padding: 10px 10px; font-size: 12px; font-weight: 600; color: #4e5969; text-align: left; white-space: nowrap; position: sticky; top: 0; z-index: 1; border-bottom: 2px solid #e5e6eb; }}
tbody td {{ padding: 8px 10px; font-size: 13px; border-bottom: 1px solid #f2f3f5; vertical-align: top; word-break: break-all; }}
tbody tr:hover {{ background: #f7f8fa; }}
tbody tr:nth-child(even) {{ background: #fafbfc; }}
tbody tr:nth-child(even):hover {{ background: #f2f3f5; }}
.col-idx {{ width: 1%; text-align: center; color: #86909c; white-space: nowrap; }}
.col-layer, .col-taskType {{ width: 1%; white-space: nowrap; }}
td.col-taskType {{ white-space: nowrap; }}
.col-module {{ width: 1%; white-space: nowrap; }}
.col-testType {{ width: 1%; white-space: nowrap; }}
.col-testTask {{ width: 1%; white-space: nowrap; }}
.col-id {{ width: 1%; white-space: nowrap; }}
.col-name {{ width: 15%; }}
.col-score {{ width: 1%; text-align: center; white-space: nowrap; font-size: 12px; }}
.col-actual {{ width: 10%; min-width: 80px; }}
.col-judge {{ width: 160px; white-space: nowrap; }}
.tag {{ display: inline-block; padding: 1px 8px; border-radius: 3px; font-size: 12px; }}
.tag-must {{ background: #e8f3ff; color: #1a73e8; }}
.tag-optional {{ background: #fff7e8; color: #ff7d00; }}
.tag-redline {{ background: #ffe8e8; color: #f53f3f; }}
.tag-L1 {{ background: #e8f7e8; color: #00b42a; }}
.tag-L2 {{ background: #f0e8ff; color: #722ed1; }}
.tag-L3 {{ background: #ffe8e8; color: #f53f3f; }}
.actions {{ display: flex; gap: 4px; white-space: nowrap; }}
.ms-wrap {{ position: relative; display: inline-block; }}
.ms-btn {{ height: 32px; border: 1px solid #c9cdd4; border-radius: 4px; padding: 0 8px; font-size: 13px; cursor: pointer; background: #fff; line-height: 32px; white-space: nowrap; user-select: none; }}
.ms-btn:hover {{ border-color: #1a73e8; }}
.ms-panel {{ display: none; position: absolute; top: 34px; left: 0; background: #fff; border: 1px solid #c9cdd4; border-radius: 4px; box-shadow: 0 4px 12px rgba(0,0,0,.15); z-index: 50; min-width: 200px; max-height: 300px; overflow-y: auto; padding: 6px 0; }}
.ms-panel.open {{ display: block; }}
.ms-check {{ display: flex; align-items: center; gap: 6px; padding: 4px 12px; font-size: 13px; cursor: pointer; white-space: nowrap; }}
.ms-check:hover {{ background: #f2f3f5; }}
.ms-check input {{ margin: 0; }}
.actions .btn {{ height: 26px; padding: 0 8px; font-size: 12px; }}
.btn-edit {{ background: #00b42a; color: #fff; }}
.btn-edit:hover {{ background: #009a25; }}
.btn-del {{ background: #ff7d00; color: #fff; }}
.btn-del:hover {{ background: #e66a00; }}

.modal-overlay {{ display: none; position: fixed; inset: 0; background: rgba(0,0,0,.45); z-index: 100; justify-content: center; align-items: flex-start; padding-top: 60px; }}
.modal-overlay.active {{ display: flex; }}
.modal {{ background: #fff; border-radius: 8px; width: 800px; max-width: 95vw; max-height: 85vh; overflow-y: auto; box-shadow: 0 8px 40px rgba(0,0,0,.2); }}
.modal-header {{ padding: 16px 24px; border-bottom: 1px solid #e5e6eb; display: flex; justify-content: space-between; align-items: center; }}
.modal-header h3 {{ font-size: 16px; }}
.modal-close {{ background: none; border: none; font-size: 20px; cursor: pointer; color: #86909c; }}
.modal-body {{ padding: 20px 24px; }}
.form-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }}
.form-group {{ display: flex; flex-direction: column; gap: 4px; }}
.form-group.full {{ grid-column: 1 / -1; }}
.form-group label {{ font-size: 13px; color: #4e5969; font-weight: 500; }}
.form-group input, .form-group select {{ height: 34px; border: 1px solid #c9cdd4; border-radius: 4px; padding: 0 10px; font-size: 13px; outline: none; }}
.form-group input:focus, .form-group select:focus {{ border-color: #1a73e8; box-shadow: 0 0 0 2px rgba(26,115,232,.15); }}
.form-group textarea {{ border: 1px solid #c9cdd4; border-radius: 4px; padding: 8px 10px; font-size: 13px; outline: none; resize: vertical; min-height: 60px; }}
.form-group textarea:focus {{ border-color: #1a73e8; box-shadow: 0 0 0 2px rgba(26,115,232,.15); }}
.modal-footer {{ padding: 12px 24px; border-top: 1px solid #e5e6eb; display: flex; justify-content: flex-end; gap: 8px; }}

.tabs {{ display: flex; background: #fff; padding: 0 24px; gap: 0; border-bottom: 1px solid #e5e6eb; }}
.tab {{ padding: 10px 20px; font-size: 14px; color: #4e5969; cursor: pointer; border-bottom: 2px solid transparent; }}
.tab.active {{ color: #1a73e8; border-bottom-color: #1a73e8; font-weight: 500; }}
.tab:hover {{ color: #1a73e8; }}
.tab-content {{ display: none; }}
.tab-content.active {{ display: block; }}

.score-cards {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 12px; margin: 20px 24px; }}
.score-card {{ padding: 16px; border-radius: 10px; color: #fff; text-align: center; }}
.score-card h4 {{ font-size: 13px; opacity: .9; margin-bottom: 6px; }}
.score-card .val {{ font-size: 26px; font-weight: 700; }}
.score-card .lbl {{ font-size: 11px; opacity: .8; margin-top: 4px; }}
.score-badge {{ display: inline-block; padding: 2px 10px; border-radius: 12px; font-size: 11px; font-weight: 700; }}
.score-s {{ background: #d4edda; color: #155724; }}
.score-a {{ background: #cce5ff; color: #004085; }}
.score-b {{ background: #fff3cd; color: #856404; }}
.score-c {{ background: #ffe5d0; color: #854d0e; }}
.score-d {{ background: #f8d7da; color: #721c24; }}
.process-cell {{ max-width: 400px; white-space: pre-wrap; font-size: 11px; color: #495057; line-height: 1.4; }}
.sub-tabs {{ display: flex; gap: 0; border-bottom: 1px solid #e5e6eb; padding: 0 24px; flex-wrap: wrap; }}
.sub-tab {{ padding: 8px 14px; font-size: 12px; color: #4e5969; cursor: pointer; border-bottom: 2px solid transparent; }}
.sub-tab.active {{ color: #1a73e8; border-bottom-color: #1a73e8; font-weight: 500; }}
.sub-tab:hover {{ color: #1a73e8; }}
.sub-content {{ display: none; padding: 16px 24px; }}
.sub-content.active {{ display: block; }}
.score-tbl {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
.score-tbl th {{ background: #667eea; color: #fff; padding: 8px 10px; text-align: left; font-weight: 500; white-space: nowrap; position: sticky; top: 0; z-index: 1; }}
.score-tbl td {{ padding: 7px 10px; border-bottom: 1px solid #f2f3f5; }}
.score-tbl tr:hover {{ background: #f7f8fa; }}

</style>
</head>
<body>

<div class="header">
  <h1>{project_name} - 业务线测试任务总清单</h1>
  <div class="stats" id="statsInfo"></div>
</div>

<div class="tabs">
  <div class="tab active" onclick="switchTab('tasks')">测试任务</div>
  <div class="tab" onclick="switchTab('weights')">权重定义</div>
  <div style="flex:1"></div>
  <div class="tab" style="color:#667eea;font-weight:600" onclick="switchTab('score')">评分计算</div>
</div>


<div class="report-bar" id="reportBar">
  <label>测试报告URL</label>
  <input type="text" id="reportUrl" placeholder="输入测试报告URL 或 选择本地文件..." style="flex:1;min-width:1200px;height:32px">
  <button class="btn btn-outline" onclick="document.getElementById('reportFile').click()">选择本地文件</button>
  <input type="file" id="reportFile" accept=".html,.htm" style="display:none" onchange="loadLocalReport(this)">
  <button class="btn btn-primary" onclick="fetchReport()">获取测试结果</button>
  <select id="reportDut" style="display:none;min-width:100px;height:32px" onchange="matchAndFill()"></select>
  <span class="report-msg" id="reportMsg"></span>
</div>
<div class="tab-content active" id="tab-tasks">
  <div class="toolbar">
    <label>模块</label>
    <select id="filterModule" onchange="render()"><option value="">全部</option></select>
    <label>测试类型</label>
    <select id="filterTestType" onchange="render()"><option value="">全部</option></select>
    <label>测试任务</label>
    <select id="filterTestTask" onchange="render()"><option value="">全部</option></select>
    <label>任务类型</label>
    <select id="filterTaskType" onchange="render()"><option value="">全部</option></select>
    <label>任务ID</label>
    <div class="ms-wrap" id="taskIdWrap">
      <div class="ms-btn" onclick="toggleMsPanel('taskIdPanel')">全部任务ID ▾</div>
      <div class="ms-panel" id="taskIdPanel">
        <label class="ms-check"><input type="checkbox" class="ms-all" onchange="msToggleAll('taskIdPanel',this)"> 全选</label>
        <div class="ms-list" id="taskIdList"></div>
      </div>
    </div>
    <div class="spacer"></div>
    <button class="btn btn-outline" onclick="clearFilters()">清除筛选</button>
    <button class="btn btn-success" onclick="exportExcel()">导出Excel</button>
    <button class="btn btn-primary" onclick="openAddModal()">新增任务</button>
  </div>

  <div class="table-wrap">
    <div class="table-info" id="tableInfo"></div>
    <div style="overflow-x:auto; max-height: calc(100vh - 260px);">
      <table>
        <thead>
          <tr>
            <th class="col-idx">#</th>
            <th class="col-layer">所属业务层</th>
            <th class="col-module">模块</th>
            <th class="col-testType">测试类型</th>
            <th class="col-testTask">测试任务</th>
            <th class="col-taskType">任务类型</th>
            <th class="col-id">任务ID</th>
            <th class="col-name">任务名称</th>
            <th>测试场景描述</th>
            <th>任务执行人</th>
            <th class="col-score">5.0分</th>
            <th class="col-score">4.0分</th>
            <th class="col-score">3.0分</th>
            <th class="col-score">2.0分</th>
            <th class="col-score">1.0分</th>
            <th class="col-score">0.0分</th>
            <th class="col-actual">测试结果</th>
            <th class="col-judge">判断机制标识</th>
            <th style="width:100px">操作</th>
          </tr>
        </thead>
        <tbody id="tableBody"></tbody>
      </table>
    </div>
  </div>
</div>

<div class="tab-content" id="tab-weights">
  <div style="padding: 24px; display: grid; grid-template-columns: 1fr 1fr; gap: 20px;">
    <div style="background:#fff;border-radius:6px;padding:20px;box-shadow:0 1px 4px rgba(0,0,0,.08);">
      <h3 style="font-size:15px;margin-bottom:12px;color:#1d2129;">任务类型权重</h3>
      <table style="box-shadow:none;"><thead><tr><th>任务类型</th><th>权重</th></tr></thead><tbody id="wt-taskType"></tbody></table>
    </div>
    <div style="background:#fff;border-radius:6px;padding:20px;box-shadow:0 1px 4px rgba(0,0,0,.08);">
      <h3 style="font-size:15px;margin-bottom:12px;color:#1d2129;">测试类型权重</h3>
      <table style="box-shadow:none;"><thead><tr><th>测试类型</th><th>权重</th></tr></thead><tbody id="wt-testType"></tbody></table>
    </div>
    <div style="background:#fff;border-radius:6px;padding:20px;box-shadow:0 1px 4px rgba(0,0,0,.08);">
      <h3 style="font-size:15px;margin-bottom:12px;color:#1d2129;">模块权重</h3>
      <table style="box-shadow:none;"><thead><tr><th>模块</th><th>权重</th></tr></thead><tbody id="wt-module"></tbody></table>
    </div>
    <div style="background:#fff;border-radius:6px;padding:20px;box-shadow:0 1px 4px rgba(0,0,0,.08);">
      <h3 style="font-size:15px;margin-bottom:12px;color:#1d2129;">业务层权重</h3>
      <table style="box-shadow:none;"><thead><tr><th>业务层</th><th>权重</th></tr></thead><tbody id="wt-layer"></tbody></table>
    </div>
  </div>
</div>


<div class="tab-content" id="tab-score">
  <div class="score-cards" id="scoreCards"></div>
  <div class="sub-tabs" id="scoreSubTabs">
    <div class="sub-tab active" onclick="switchSubTab(this,'sub-redline')">红线项</div>
    <div class="sub-tab" onclick="switchSubTab(this,'sub-task')">任务级</div>
    <div class="sub-tab" onclick="switchSubTab(this,'sub-testtype')">测试类型</div>
    <div class="sub-tab" onclick="switchSubTab(this,'sub-module')">模块级</div>
    <div class="sub-tab" onclick="switchSubTab(this,'sub-layer')">层级</div>
    <div class="sub-tab" onclick="switchSubTab(this,'sub-total')">综合</div>
    <div class="sub-tab" onclick="switchSubTab(this,'sub-stats')">统计</div>
    <div class="sub-tab" onclick="switchSubTab(this,'sub-missing')">缺失</div>
  </div>
  <div class="sub-content active" id="sub-redline"><div style="overflow-x:auto;max-height:500px;overflow-y:auto"><div id="stbl-redline"></div></div></div>
  <div class="sub-content" id="sub-task"><div style="overflow-x:auto;max-height:500px;overflow-y:auto"><div id="stbl-task"></div></div></div>
  <div class="sub-content" id="sub-testtype"><div style="overflow-x:auto;max-height:500px;overflow-y:auto"><div id="stbl-testtype"></div></div></div>
  <div class="sub-content" id="sub-module"><div style="overflow-x:auto;max-height:500px;overflow-y:auto"><div id="stbl-module"></div></div></div>
  <div class="sub-content" id="sub-layer"><div style="overflow-x:auto;max-height:500px;overflow-y:auto"><div id="stbl-layer"></div></div></div>
  <div class="sub-content" id="sub-total"><div style="overflow-x:auto;max-height:300px;overflow-y:auto"><div id="stbl-total"></div></div></div>
  <div class="sub-content" id="sub-stats"><div style="overflow-x:auto;max-height:400px;overflow-y:auto"><div id="stbl-stats"></div></div></div>
  <div class="sub-content" id="sub-missing"><div style="overflow-x:auto;max-height:400px;overflow-y:auto"><div id="stbl-missing"></div></div></div>
</div>

<div class="modal-overlay" id="modal">
  <div class="modal">
    <div class="modal-header">
      <h3 id="modalTitle">编辑任务</h3>
      <button class="modal-close" onclick="closeModal()">&times;</button>
    </div>
    <div class="modal-body">
      <div class="form-grid">
        <div class="form-group"><label>所属业务层</label><select id="f-layer"><option>L1</option><option>L2</option><option>L3</option></select></div>
        <div class="form-group"><label>模块</label><input id="f-module" list="dl-module"><datalist id="dl-module"></datalist></div>
        <div class="form-group"><label>测试类型</label><select id="f-testType"><option>性能测试</option><option>功能测试</option><option>主观体验测试</option></select></div>
        <div class="form-group"><label>测试任务</label><input id="f-testTask" list="dl-testTask"><datalist id="dl-testTask"></datalist></div>
        <div class="form-group"><label>任务类型</label><select id="f-taskType"><option>必测项</option><option>选测项</option><option>红线项</option></select></div>
        <div class="form-group"><label>任务ID</label><input id="f-taskId"></div>
        <div class="form-group full"><label>任务名称</label><textarea id="f-taskName"></textarea></div>
        <div class="form-group"><label>任务执行人</label><input id="f-executor"></div>
        <div class="form-group"><label>合同指标</label><input id="f-contractTarget"></input></div>
        <div class="form-group"><label>5.0分</label><input id="f-score5"></div>
        <div class="form-group"><label>4.0分</label><input id="f-score4"></div>
        <div class="form-group"><label>3.0分</label><input id="f-score3"></div>
        <div class="form-group"><label>2.0分</label><input id="f-score2"></div>
        <div class="form-group"><label>1.0分</label><input id="f-score1"></div>
        <div class="form-group"><label>0.0分</label><input id="f-score0"></div>
        <div class="form-group"><label>测试结果</label><input id="f-actualResult"></div>
        <div class="form-group"><label>判断机制标识</label><select id="f-judgeMode"><option>正常评分（Normal）</option><option>锁定上限（Cap）</option><option>特批降级（Waived）</option><option>红线置零（Veto）</option><option>离散评分（SoftVeto）</option></select></div>
        <div class="form-group"><label>竞品指标</label><input id="f-competeTarget"></div>
        <div class="form-group"><label>KA实际交付指标</label><input id="f-kaTarget"></div>
        <div class="form-group full"><label>测试场景描述</label><textarea id="f-sceneDesc"></textarea></div>
        <div class="form-group full"><label>测试集构成概述</label><textarea id="f-testSetSummary"></textarea></div>
        <div class="form-group"><label>测试环境</label><input id="f-testEnv"></div>
        <div class="form-group"><label>测试工具</label><input id="f-testTool"></div>
        <div class="form-group full"><label>统计指标及计算公式</label><input id="f-metricFormula"></div>
      </div>
    </div>
    <div class="modal-footer">
      <button class="btn btn-outline" onclick="closeModal()">取消</button>
      <button class="btn btn-primary" onclick="saveTask()">保存</button>
    </div>
  </div>
</div>

<script>
const INIT_DATA = {json_data};
const PROJECT_NAME = "{project_name}";

let tasks = JSON.parse(JSON.stringify(INIT_DATA.tasks));
const weights = INIT_DATA.weights;
let editIndex = -1;

function unique(arr, key) {{ return [...new Set(arr.map(t => t[key]))].filter(Boolean).sort(); }}
function esc(s) {{ if (!s) return ''; return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;'); }}

function toggleMsPanel(panelId) {{
  const panel = document.getElementById(panelId);
  panel.classList.toggle('open');
  if (panel.classList.contains('open')) {{
    const handler = (e) => {{
      if (!panel.parentElement.contains(e.target)) {{ panel.classList.remove('open'); document.removeEventListener('click', handler); }}
    }};
    setTimeout(() => document.addEventListener('click', handler), 0);
  }}
}}
function msUpdateBtn(panelId) {{
  const panel = document.getElementById(panelId);
  const btn = panel.previousElementSibling;
  const checked = panel.querySelectorAll('.ms-item:checked');
  btn.textContent = checked.length ? '已选 ' + checked.length + ' 项 ▾' : '全部任务ID ▾';
}}
function msToggleAll(panelId, el) {{
  const panel = document.getElementById(panelId);
  panel.querySelectorAll('.ms-item').forEach(cb => cb.checked = el.checked);
  msUpdateBtn(panelId); render();
}}
function getMsSelected(listId) {{
  return [...document.getElementById(listId).querySelectorAll('.ms-item:checked')].map(cb => cb.value);
}}
function fillMsCheckboxes(listId, items) {{
  const el = document.getElementById(listId);
  const prev = new Set(getMsSelected(listId));
  el.innerHTML = items.map(i => '<label class="ms-check"><input type="checkbox" class="ms-item" value="'+esc(i)+'" '+(prev.has(i)?'checked':'')+' onchange="msUpdateBtn(this.closest(\\'.ms-panel\\').id); render()"> '+esc(i)+'</label>').join('');
  const panel = el.closest('.ms-panel');
  const allCb = panel.querySelector('.ms-all');
  allCb.checked = items.length > 0 && items.every(i => prev.has(i));
  msUpdateBtn(panel.id);
}}
function msClearAll(listId) {{
  const el = document.getElementById(listId);
  el.querySelectorAll('.ms-item').forEach(cb => cb.checked = false);
  el.closest('.ms-panel').querySelector('.ms-all').checked = false;
}}

function initFilters() {{
  const modules = unique(tasks, 'module');
  const testTypes = unique(tasks, 'testType');
  const testTasks = unique(tasks, 'testTask');
  fillSelect('filterModule', modules);
  fillSelect('filterTestType', testTypes);
  fillSelect('filterTestTask', testTasks);
  fillSelect('filterTaskType', ['必测项', '选测项', '红线项']);
  fillMsCheckboxes('taskIdList', unique(tasks, 'taskId'));
  fillDatalist('dl-module', modules);
  fillDatalist('dl-testTask', testTasks);
}}

function fillSelect(id, items) {{
  const el = document.getElementById(id);
  const val = el.value;
  el.innerHTML = '<option value="">全部</option>' + items.map(i => '<option value="'+esc(i)+'">'+esc(i)+'</option>').join('');
  el.value = val;
}}

function fillDatalist(id, items) {{
  document.getElementById(id).innerHTML = items.map(i => '<option value="'+esc(i)+'">').join('');
}}

function getFiltered() {{
  const fModule = document.getElementById('filterModule').value;
  const fTestType = document.getElementById('filterTestType').value;
  const fTestTask = document.getElementById('filterTestTask').value;
  const fTaskType = document.getElementById('filterTaskType').value;
  const fTaskIds = getMsSelected('taskIdList');
  return tasks.filter(t => {{
    if (fModule && t.module !== fModule) return false;
    if (fTestType && t.testType !== fTestType) return false;
    if (fTestTask && t.testTask !== fTestTask) return false;
    if (fTaskType && t.taskType !== fTaskType) return false;
    if (fTaskIds.length && !fTaskIds.includes(t.taskId)) return false;
    return true;
  }});
}}

function updateTaskIdOptions() {{
  const fModule = document.getElementById('filterModule').value;
  const fTestType = document.getElementById('filterTestType').value;
  const fTestTask = document.getElementById('filterTestTask').value;
  const fTaskType = document.getElementById('filterTaskType').value;
  const candidates = tasks.filter(t => {{
    if (fModule && t.module !== fModule) return false;
    if (fTestType && t.testType !== fTestType) return false;
    if (fTestTask && t.testTask !== fTestTask) return false;
    if (fTaskType && t.taskType !== fTaskType) return false;
    return true;
  }});
  fillMsCheckboxes('taskIdList', unique(candidates, 'taskId'));
}}

function render() {{
  updateTaskIdOptions();
  const filtered = getFiltered();
  const tbody = document.getElementById('tableBody');
  const mustCount = filtered.filter(t => t.taskType === '必测项').length;
  const optionalCount = filtered.filter(t => t.taskType === '选测项').length;
  document.getElementById('tableInfo').textContent = '共 '+filtered.length+' 条（必测 '+mustCount+' 条，选测 '+optionalCount+' 条）';
  tbody.innerHTML = filtered.map((t, i) => {{
    const idx = tasks.indexOf(t);
    return '<tr>'
      +'<td class="col-idx">'+(i+1)+'</td>'
      +'<td><span class="tag tag-'+t.layer+'">'+esc(t.layer)+'</span></td>'
      +'<td style="white-space:nowrap">'+esc(t.module)+'</td>'
      +'<td style="white-space:nowrap">'+esc(t.testType)+'</td>'
      +'<td style="white-space:nowrap">'+esc(t.testTask)+'</td>'
      +'<td style="white-space:nowrap"><span class="tag '+(t.taskType==='红线项'?'tag-redline':t.taskType==='必测项'?'tag-must':'tag-optional')+'">'+esc(t.taskType)+'</span></td>'
      +'<td style="white-space:nowrap">'+esc(t.taskId)+'</td>'
      +'<td class="col-name">'+esc(t.taskName)+'</td>'
      +'<td>'+esc(t.sceneDesc)+'</td>'
      +'<td>'+esc(t.executor)+'</td>'
      +'<td class="col-score">'+esc(t.score5)+'</td>'
      +'<td class="col-score">'+esc(t.score4)+'</td>'
      +'<td class="col-score">'+esc(t.score3)+'</td>'
      +'<td class="col-score">'+esc(t.score2)+'</td>'
      +'<td class="col-score">'+esc(t.score1)+'</td>'
      +'<td class="col-score">'+esc(t.score0)+'</td>'
      +'<td class="col-actual" contenteditable="true" onblur="tasks['+idx+'].actualResult=this.textContent">'+esc(t.actualResult)+'</td>'
      +'<td class="col-judge"><select onchange="tasks['+idx+'].judgeMode=this.value" style="width:100%;border:none;background:transparent;font-size:12px;outline:none;cursor:pointer"><option value="正常评分（Normal）" '+(t.judgeMode==='正常评分（Normal）'?'selected':'')+'>正常评分（Normal）</option><option value="锁定上限（Cap）" '+(t.judgeMode==='锁定上限（Cap）'?'selected':'')+'>锁定上限（Cap）</option><option value="特批降级（Waived）" '+(t.judgeMode==='特批降级（Waived）'?'selected':'')+'>特批降级（Waived）</option><option value="红线置零（Veto）" '+(t.judgeMode==='红线置零（Veto）'?'selected':'')+'>红线置零（Veto）</option><option value="离散评分（SoftVeto）" '+(t.judgeMode==='离散评分（SoftVeto）'?'selected':'')+'>离散评分（SoftVeto）</option></select></td>'
      +'<td style="white-space:nowrap"><div class="actions">'
      +'<button class="btn btn-edit" onclick="openEditModal('+idx+')">编辑</button>'
      +'<button class="btn btn-del" onclick="deleteTask('+idx+')">删除</button>'
      +'</div></td></tr>';
  }}).join('');
  document.getElementById('statsInfo').textContent = '共 '+tasks.length+' 条任务 | 必测 '+tasks.filter(t=>t.taskType==='必测项').length+' 条 | 选测 '+tasks.filter(t=>t.taskType==='选测项').length+' 条 | 红线 '+tasks.filter(t=>t.taskType==='红线项').length+' 条';
}}

function switchTab(name) {{
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  if (name==='tasks') document.querySelectorAll('.tab')[0].classList.add('active');
  else if (name==='weights') document.querySelectorAll('.tab')[1].classList.add('active');
  else if (name==='score') {{ document.querySelectorAll('.tab')[2].classList.add('active'); performScoreCalc(); }}
  document.getElementById('tab-tasks').classList.toggle('active', name==='tasks');
  document.getElementById('tab-weights').classList.toggle('active', name==='weights');
  document.getElementById('tab-score').classList.toggle('active', name==='score');
}}

function renderWeights() {{
  ['taskType','testType','module','layer'].forEach(key => {{
    const tbody = document.getElementById('wt-'+key);
    tbody.innerHTML = weights[key].map((w, i) => '<tr><td>'+esc(w.name)+'</td><td><input type="number" step="0.01" value="'+w.weight+'" style="width:80px;height:28px;border:1px solid #c9cdd4;border-radius:3px;padding:0 6px;font-size:13px;text-align:center;outline:none" onchange="weights[\\''+key+'\\']['+i+'].weight=parseFloat(this.value)||0"></td></tr>').join('');
  }});
}}

function openAddModal() {{ editIndex = -1; document.getElementById('modalTitle').textContent = '新增任务'; clearForm(); document.getElementById('modal').classList.add('active'); }}
function openEditModal(idx) {{
  editIndex = idx; document.getElementById('modalTitle').textContent = '编辑任务';
  const t = tasks[idx];
  document.getElementById('f-layer').value = t.layer;
  document.getElementById('f-module').value = t.module;
  document.getElementById('f-testType').value = t.testType;
  document.getElementById('f-testTask').value = t.testTask;
  document.getElementById('f-taskType').value = t.taskType;
  document.getElementById('f-taskId').value = t.taskId;
  document.getElementById('f-taskName').value = t.taskName;
  document.getElementById('f-executor').value = t.executor;
  document.getElementById('f-score5').value = t.score5;
  document.getElementById('f-score4').value = t.score4;
  document.getElementById('f-score3').value = t.score3;
  document.getElementById('f-score2').value = t.score2;
  document.getElementById('f-score1').value = t.score1;
  document.getElementById('f-score0').value = t.score0;
  document.getElementById('f-actualResult').value = t.actualResult;
  document.getElementById('f-judgeMode').value = t.judgeMode;
  document.getElementById('f-contractTarget').value = t.contractTarget;
  document.getElementById('f-competeTarget').value = t.competeTarget;
  document.getElementById('f-kaTarget').value = t.kaTarget;
  document.getElementById('f-sceneDesc').value = t.sceneDesc;
  document.getElementById('f-testSetSummary').value = t.testSetSummary;
  document.getElementById('f-testEnv').value = t.testEnv;
  document.getElementById('f-testTool').value = t.testTool;
  document.getElementById('f-metricFormula').value = t.metricFormula;
  document.getElementById('modal').classList.add('active');
}}
function closeModal() {{ document.getElementById('modal').classList.remove('active'); }}
function clearForm() {{
  ['f-layer','f-module','f-testType','f-testTask','f-taskType','f-taskId','f-taskName','f-executor','f-score5','f-score4','f-score3','f-score2','f-score1','f-score0','f-actualResult','f-judgeMode','f-contractTarget','f-competeTarget','f-kaTarget','f-sceneDesc','f-testSetSummary','f-testEnv','f-testTool','f-metricFormula'].forEach(id => {{ document.getElementById(id).value = ''; }});
  document.getElementById('f-layer').value = 'L1';
  document.getElementById('f-testType').value = '性能测试';
  document.getElementById('f-taskType').value = '选测项';
  document.getElementById('f-executor').value = '业务线质量';
}}

function saveTask() {{
  const obj = {{
    layer: document.getElementById('f-layer').value,
    module: document.getElementById('f-module').value,
    testType: document.getElementById('f-testType').value,
    testTask: document.getElementById('f-testTask').value,
    taskType: document.getElementById('f-taskType').value,
    taskId: document.getElementById('f-taskId').value,
    taskName: document.getElementById('f-taskName').value,
    executor: document.getElementById('f-executor').value,
    sceneDesc: document.getElementById('f-sceneDesc').value,
    testSetSummary: document.getElementById('f-testSetSummary').value,
    testEnv: document.getElementById('f-testEnv').value,
    testTool: document.getElementById('f-testTool').value,
    metricFormula: document.getElementById('f-metricFormula').value,
    contractTarget: document.getElementById('f-contractTarget').value,
    competeTarget: document.getElementById('f-competeTarget').value,
    kaTarget: document.getElementById('f-kaTarget').value,
    score5: document.getElementById('f-score5').value,
    score4: document.getElementById('f-score4').value,
    score3: document.getElementById('f-score3').value,
    score2: document.getElementById('f-score2').value,
    score1: document.getElementById('f-score1').value,
    score0: document.getElementById('f-score0').value,
    actualResult: document.getElementById('f-actualResult').value,
    judgeMode: document.getElementById('f-judgeMode').value,
  }};
  if (!obj.taskId || !obj.taskName) {{ alert('任务ID和任务名称不能为空'); return; }}
  if (editIndex >= 0) tasks[editIndex] = obj; else tasks.push(obj);
  closeModal(); initFilters(); render();
}}

function deleteTask(idx) {{
  if (!confirm('确定删除任务 "'+tasks[idx].taskId+'" 吗？')) return;
  tasks.splice(idx, 1); initFilters(); render();
}}

function addWeightSheets(wb) {{
  const wtTaskType = weights['taskType'];
  const wtTestType = weights['testType'];
  const maxWtRows = Math.max(wtTaskType.length, wtTestType.length);
  const wtData1 = [['单个任务类型','任务权重','','','测试类型','测试类型权重']];
  for (let i = 0; i < maxWtRows; i++) {{
    const r = ['','','','','',''];
    if (i < wtTaskType.length) {{ r[0] = wtTaskType[i].name; r[1] = wtTaskType[i].weight; }}
    if (i < wtTestType.length) {{ r[4] = wtTestType[i].name; r[5] = wtTestType[i].weight; }}
    wtData1.push(r);
  }}
  XLSX.utils.book_append_sheet(wb, XLSX.utils.aoa_to_sheet(wtData1), '任务权重定义');
  ['module','layer'].forEach(key => {{
    const titleMap = {{module:'模块权重定义',layer:'业务层权重定义'}};
    const headerMap = {{module:['模块','模块权重'],layer:['所属业务层','业务层权重']}};
    const wData = [headerMap[key], ...weights[key].map(w => [w.name, w.weight])];
    XLSX.utils.book_append_sheet(wb, XLSX.utils.aoa_to_sheet(wData), titleMap[key]);
  }});
}}

function exportExcel() {{
  const filtered = getFiltered();

  // === 文件1: 评分总表（全量，不含测试结果/判断机制标识）===
  const wb1 = XLSX.utils.book_new();
  const scoreHeaders = ['所属业务层','模块','测试类型','测试任务','任务类型','任务ID','任务名称','任务执行人','测试场景描述','测试集构成概述','测试环境','测试工具','统计指标及计算公式','合同指标','竞品指标','KA实际交付指标','单个任务原始评分：对应的任务测试性能结果区间',''];
  const scoreHeaders2 = ['','','','','','','','','','','','','','','','','5.0分','4.0分','3.0分','2.0分','1.0分','0.0分'];
  const scoreKeys = ['layer','module','testType','testTask','taskType','taskId','taskName','executor','sceneDesc','testSetSummary','testEnv','testTool','metricFormula','contractTarget','competeTarget','kaTarget','score5','score4','score3','score2','score1','score0'];
  const scoreData = [scoreHeaders, scoreHeaders2, ...tasks.map(t => scoreKeys.map(k => t[k] || ''))];
  const ws1 = XLSX.utils.aoa_to_sheet(scoreData);
  ws1['!merges'] = [];
  for (let c = 0; c < 16; c++) ws1['!merges'].push({{s:{{r:0,c:c}},e:{{r:1,c:c}}}});
  ws1['!merges'].push({{s:{{r:0,c:16}},e:{{r:0,c:21}}}});
  ws1['!cols'] = scoreKeys.map((_, i) => {{ if (i===6) return {{wch:60}}; if (i>=16) return {{wch:10}}; return {{wch:16}}; }});
  XLSX.utils.book_append_sheet(wb1, ws1, '产品线全量测试任务表');
  addWeightSheets(wb1);
  XLSX.writeFile(wb1, PROJECT_NAME + '-评分总表.xlsx');

  // === 文件2: 测试范围（筛选结果）===
  const wb2 = XLSX.utils.book_new();
  const scopeHeaders = ['所属业务层','模块','测试类型','测试任务','任务类型','任务ID','任务名称','任务执行人'];
  const scopeKeys = ['layer','module','testType','testTask','taskType','taskId','taskName','executor'];
  const scopeData = [scopeHeaders, ...filtered.map(t => scopeKeys.map(k => t[k] || ''))];
  const ws2 = XLSX.utils.aoa_to_sheet(scopeData);
  ws2['!cols'] = scopeKeys.map((_, i) => {{ if (i===6) return {{wch:60}}; return {{wch:16}}; }});
  XLSX.utils.book_append_sheet(wb2, ws2, '测试范围');
  XLSX.writeFile(wb2, PROJECT_NAME + '-测试范围.xlsx');

  // === 文件3: 项目实测数据（筛选结果）===
  const wb3 = XLSX.utils.book_new();
  const resultHeaders = ['所属业务层','模块','测试类型','测试任务','任务类型','任务ID','任务名称','测试结果','判断机制标识'];
  const resultKeys = ['layer','module','testType','testTask','taskType','taskId','taskName','actualResult','judgeMode'];
  const resultData = [resultHeaders, ...filtered.map(t => resultKeys.map(k => t[k] || ''))];
  const ws3 = XLSX.utils.aoa_to_sheet(resultData);
  ws3['!cols'] = resultKeys.map((_, i) => {{ if (i===6) return {{wch:60}}; return {{wch:16}}; }});
  XLSX.utils.book_append_sheet(wb3, ws3, '项目实测数据');
  XLSX.writeFile(wb3, PROJECT_NAME + '-项目实测数据.xlsx');

  alert('已导出 3 个文件：\\n1. ' + PROJECT_NAME + '-评分总表.xlsx\\n2. ' + PROJECT_NAME + '-测试范围.xlsx\\n3. ' + PROJECT_NAME + '-项目实测数据.xlsx');
}}

function clearFilters() {{
  document.getElementById('filterModule').value = '';
  document.getElementById('filterTestType').value = '';
  document.getElementById('filterTestTask').value = '';
  document.getElementById('filterTaskType').value = '';
  msClearAll('taskIdList'); msUpdateBtn('taskIdPanel');
  render();
}}


// ========== 评分计算引擎 ==========
function switchSubTab(el, id) {{
  el.parentElement.querySelectorAll('.sub-tab').forEach(t => t.classList.remove('active'));
  el.classList.add('active');
  el.parentElement.parentElement.querySelectorAll('.sub-content').forEach(c => c.classList.remove('active'));
  document.getElementById(id).classList.add('active');
}}

function parseResult(value, isPct) {{
  if (value===null||value===undefined||value==='') return NaN;
  if (typeof value==='number') return isPct&&value>0&&value<=1.0 ? value*100 : value;
  value = String(value).trim();
  if (value===''||value==='NaN') return NaN;
  if (value.includes('%')) return parseFloat(value.match(/[\d\.]+/)[0]);
  if (value.includes('次')) return parseInt(value.match(/\d+/)[0]);
  if (value.toLowerCase().includes('ms')) return parseFloat(value.match(/[\d\.]+/)[0]);
  if (value.toLowerCase().includes('s')&&!value.toLowerCase().includes('ms')) return parseFloat(value.match(/[\d\.]+/)[0]);
  if (value.includes('dB')) return parseFloat(value.match(/-?[\d\.]+/)[0]);
  if (value.includes('G')) return parseFloat(value.match(/-?[\d\.]+/)[0])*1024;
  if (value.includes('M')) return parseFloat(value.match(/-?[\d\.]+/)[0]);
  const num = parseFloat(value);
  if (!isNaN(num)) return isPct&&num>0&&num<=1.0 ? num*100 : num;
  return NaN;
}}

function parseCond(condStr, value) {{
  if (!condStr||condStr==='/'||condStr==='') return null;
  condStr = String(condStr).replace(/＜/g,'<').replace(/＞/g,'>').replace(/＝/g,'=');
  const isMem = condStr.includes('M')||condStr.includes('G');
  const isDb = condStr.includes('dB');
  function exNum(s) {{
    if (isMem) {{ const n=parseFloat(s.match(/-?[\d\.]+/)[0]); return s.includes('G')?n*1024:n; }}
    if (isDb) return parseFloat(s.match(/-?[\d\.]+/)[0]);
    return parseFloat(s.match(/[\d\.]+/)[0]);
  }}
  if (condStr.includes('&')) {{
    const parts = condStr.split('&');
    let lo=null,hi=null,loEq=false,hiEq=false;
    for (let p of parts) {{
      p=p.trim();
      if (p.startsWith('\u2265')) {{ lo=exNum(p); loEq=true; }}
      else if (p.startsWith('>')) {{ lo=exNum(p); loEq=false; }}
      else if (p.startsWith('\u2264')) {{ hi=exNum(p); hiEq=true; }}
      else if (p.startsWith('<')) {{ hi=exNum(p); hiEq=false; }}
    }}
    if (lo!==null&&hi!==null) return loEq&&hiEq?lo<=value&&value<=hi:loEq?lo<=value&&value<hi:hiEq?lo<value&&value<=hi:lo<value&&value<hi;
    if (lo!==null) return loEq?value>=lo:value>lo;
    if (hi!==null) return hiEq?value<=hi:value<hi;
  }}
  if (condStr.startsWith('\u2265')) return value>=exNum(condStr);
  if (condStr.startsWith('>')) return value>exNum(condStr);
  if (condStr.startsWith('<')) return value<exNum(condStr);
  if (condStr.startsWith('\u2264')) return value<=exNum(condStr);
  if (condStr.startsWith('=')) return value===exNum(condStr);
  if (condStr==='0次') return value===0;
  if (condStr.includes('次')) {{
    const n=parseInt(condStr.match(/\d+/)[0]);
    if (condStr.includes('\u2265')) return value>=n;
    if (condStr.includes('\u2264')) return value<=n;
    if (condStr.includes('>')) return value>n;
    if (condStr.includes('<')) return value<n;
    return value===n;
  }}
  return null;
}}

function getRawScore(task) {{
  const r = task.result || task.actualResult;
  if (r===null||r===undefined||r==='') return NaN;
  let isPct = false;
  for (const sc of ['score5','score4','score3','score2','score1','score0']) {{
    const cv = task[sc];
    if (cv&&String(cv).includes('%')) {{ isPct=true; break; }}
  }}
  const value = parseResult(r, isPct);
  if (isNaN(value)) return NaN;
  const scoreList = task.taskType!=='红线项' ? [5,4,3,2,1,0] : [5,0];
  for (const score of scoreList) {{
    const cond = task['score'+score];
    if (!cond||cond==='/'||cond==='') continue;
    if (parseCond(String(cond), value)) return score;
  }}
  // 有有效结果但低于所有阈值，返回最低分 0
  return 0;
}}

function extractFlag(rating) {{
  if (!rating) return 'Normal';
  rating = String(rating);
  if (rating.includes('SoftVeto')) return 'SoftVeto';
  if (rating.includes('Veto')) return 'Veto';
  if (rating.includes('Cap')) {{ if (rating.startsWith('D')) return 'D_Cap'; if (rating.startsWith('C')) return 'C_Cap'; }}
  if (rating.includes('Waived')) return 'Waived';
  return 'Normal';
}}

function getRating(score, capType, svApplied) {{
  score = Math.round(score*10000)/10000;
  if (capType==='D_Cap') return 'D（Cap）';
  if (capType==='C_Cap') return 'C（Cap）';
  let r;
  if (score>=4.5) r='S'; else if (score>=3.5) r='A'; else if (score>=3.0) r='B'; else if (score>=2.0) r='C'; else r='D';
  if (svApplied&&!capType&&!r.includes('（')) r+='（SoftVeto）';
  return r;
}}

function getScoreCls(rating) {{
  if (!rating) return '';
  const r = String(rating);
  if (r.startsWith('S')) return 'score-s';
  if (r.startsWith('A')) return 'score-a';
  if (r.startsWith('B')) return 'score-b';
  if (r.startsWith('C')) return 'score-c';
  if (r.startsWith('D')) return 'score-d';
  return '';
}}

const SCORE_RULES = [
  {{type:'红线项',raw:5,flag:'正常评分（Normal）',final:5,rating:'S'}},
  {{type:'红线项',raw:0,flag:'红线置零（Veto）',final:0,rating:'D（Veto）'}},
  {{type:'必测项',raw:5,flag:'正常评分（Normal）',final:5,rating:'S'}},
  {{type:'必测项',raw:4,flag:'正常评分（Normal）',final:4,rating:'A'}},
  {{type:'必测项',raw:3,flag:'离散评分（SoftVeto）',final:3,rating:'B（SoftVeto）'}},
  {{type:'必测项',raw:2,flag:'锁定上限（Cap）',final:2,rating:'C（Cap）'}},
  {{type:'必测项',raw:2,flag:'特批降级（Waived）',final:3,rating:'B（Waived）'}},
  {{type:'必测项',raw:1,flag:'锁定上限（Cap）',final:1,rating:'D（Cap）'}},
  {{type:'选测项',raw:5,flag:'正常评分（Normal）',final:5,rating:'S'}},
  {{type:'选测项',raw:4,flag:'正常评分（Normal）',final:4,rating:'A'}},
  {{type:'选测项',raw:3,flag:'正常评分（Normal）',final:3,rating:'B'}},
  {{type:'选测项',raw:2,flag:'正常评分（Normal）',final:2,rating:'C'}},
  {{type:'选测项',raw:1,flag:'正常评分（Normal）',final:1,rating:'D'}},
];

function getTaskWeight(tt) {{
  if (tt==='红线项') return 0;
  const w = weights.taskType.find(w => w.name===tt);
  return w ? w.weight : 0;
}}
function getTestTypeWeight(tt) {{
  const w = weights.testType.find(w => w.name===tt);
  return w ? w.weight : 0;
}}
function getModuleWeight(m) {{
  const w = weights.module.find(w => w.name===m);
  return w ? w.weight : 0;
}}
function getLayerWeight(l) {{
  const w = weights.layer.find(w => w.name===l);
  return w ? w.weight : 0;
}}

function mkScoreTbl(data, cols, ratingCol) {{
  if (!data||!data.length) return '<p style="padding:20px;text-align:center;color:#86909c">暂无数据</p>';
  let h = '<table class="score-tbl"><thead><tr>';
  cols.forEach(c => h+='<th>'+c+'</th>');
  h += '</tr></thead><tbody>';
  data.forEach(row => {{
    h += '<tr>';
    cols.forEach(c => {{
      let v = row[c]; if (v===null||v===undefined) v='';
      if (typeof v==='number'&&!Number.isInteger(v)) v=Math.round(v*100)/100;
      if (c==='计算过程') h+='<td class="process-cell">'+v+'</td>';
      else if (c===ratingCol) h+='<td><span class="score-badge '+getScoreCls(v)+'">'+v+'</span></td>';
      else h+='<td>'+v+'</td>';
    }});
    h += '</tr>';
  }});
  h += '</tbody></table>';
  return h;
}}

function performScoreCalc() {{
  const filtered = getFiltered();
  const merged = filtered.map(t => ({{
    layer:t.layer, module:t.module, testType:t.testType, testTask:t.testTask,
    taskType:t.taskType, taskId:t.taskId, taskName:t.taskName, executor:t.executor,
    result:t.actualResult, flag_raw:t.judgeMode||'正常评分（Normal）',
    score5:t.score5, score4:t.score4, score3:t.score3, score2:t.score2, score1:t.score1, score0:t.score0
  }}));

  // 原始评分
  merged.forEach(t => t.rawScore = getRawScore(t));

  // 最终评分&评级
  merged.forEach(t => {{
    if (isNaN(t.rawScore)) {{ t.finalScore=NaN; t.finalRating=null; return; }}
    let rule = SCORE_RULES.find(r => r.type===t.taskType&&r.raw===t.rawScore&&r.flag===t.flag_raw);
    if (!rule) rule = SCORE_RULES.find(r => r.type===t.taskType&&r.raw===t.rawScore);
    if (rule) {{ t.finalScore=rule.final; t.finalRating=rule.rating; }}
    else {{ t.finalScore=NaN; t.finalRating=null; }}
  }});

  const redlineDf = merged.filter(t => t.taskType==='红线项');
  const nonRedDf = merged.filter(t => t.taskType!=='红线项');
  const taskCols = ['所属业务层','模块','测试类型','测试任务','任务类型','任务ID','任务名称','测试结果','原始评分','最终评分','最终评级'];
  const rename = t => ({{
    '所属业务层':t.layer,'模块':t.module,'测试类型':t.testType,'测试任务':t.testTask,
    '任务类型':t.taskType,'任务ID':t.taskId,'任务名称':t.taskName,'测试结果':t.result,
    '原始评分':t.rawScore,'最终评分':t.finalScore,'最终评级':t.finalRating
  }});
  document.getElementById('stbl-redline').innerHTML = mkScoreTbl(redlineDf.map(rename), taskCols, '最终评级');
  document.getElementById('stbl-task').innerHTML = mkScoreTbl(nonRedDf.map(rename), taskCols, '最终评级');

  // 模块级计算
  const modGroups = {{}};
  nonRedDf.forEach(t => {{
    const key = t.layer+'|'+t.module;
    if (!modGroups[key]) modGroups[key]=[];
    modGroups[key].push(t);
  }});

  const ttResults = [], modResults = [];
  for (const mKey in modGroups) {{
    const [layer, mod] = mKey.split('|');
    const mTasks = modGroups[mKey].filter(t => !isNaN(t.finalScore));
    const procLines = [];

    // 红线Veto检查
    const rlVeto = merged.filter(t => t.layer===layer&&t.module===mod&&t.taskType==='红线项'&&t.result&&t.finalRating&&String(t.finalRating).includes('Veto'));
    if (rlVeto.length>0) {{
      procLines.push('红线项Veto('+rlVeto.map(t=>t.taskName).join(',')+'), 模块归零');
      const ttTypes = [...new Set(modGroups[mKey].map(t=>t.testType))];
      ttTypes.forEach(tt => ttResults.push({{'所属业务层':layer,'模块':mod,'测试类型':tt,'测试类型得分':0,'计算过程':'红线Veto'}}));
      modResults.push({{'所属业务层':layer,'模块':mod,'模块最终评分':0,'模块最终评级':'D（Veto）','计算过程':procLines.join('\\n')}});
      continue;
    }}
    if (mTasks.length===0) {{
      if (merged.filter(t=>t.layer===layer&&t.module===mod&&t.taskType==='红线项').length===0) {{
        procLines.push('无有效任务');
        modResults.push({{'所属业务层':layer,'模块':mod,'模块最终评分':0,'模块最终评级':'D','计算过程':procLines.join('\\n')}});
      }} else procLines.push('仅红线项, 不参与加权');
      continue;
    }}

    mTasks.forEach(t => t.flag = extractFlag(t.finalRating));
    const vetoT = mTasks.filter(t => t.flag==='Veto');
    if (vetoT.length>0) {{
      procLines.push('Veto任务('+vetoT.map(t=>t.taskName).join(',')+')');
      const ttTypes = [...new Set(mTasks.map(t=>t.testType))];
      ttTypes.forEach(tt => ttResults.push({{'所属业务层':layer,'模块':mod,'测试类型':tt,'测试类型得分':0,'计算过程':'Veto'}}));
      modResults.push({{'所属业务层':layer,'模块':mod,'模块最终评分':0,'模块最终评级':'D（Veto）','计算过程':procLines.join('\\n')}});
      continue;
    }}

    const hasDC = mTasks.some(t=>t.flag==='D_Cap'), hasCC = mTasks.some(t=>t.flag==='C_Cap'), hasSV = mTasks.some(t=>t.flag==='SoftVeto');
    let capLim=null, capType=null;
    if (hasDC) {{ capLim=1.9; capType='D_Cap'; procLines.push('D(Cap)→上限1.9'); }}
    else if (hasCC) {{ capLim=2.9; capType='C_Cap'; procLines.push('C(Cap)→上限2.9'); }}

    // 测试类型分组
    procLines.push('【测试类型加权】');
    const ttGrps = {{}};
    mTasks.forEach(t => {{ const tt=t.testType; if(!ttGrps[tt])ttGrps[tt]=[]; ttGrps[tt].push(t); }});
    const ttScores = {{}};
    for (const tt in ttGrps) {{
      let ttT = ttGrps[tt];
      const ttProc = ['测试类型['+tt+']：'];
      const ttVeto = ttT.filter(t=>t.flag==='Veto');
      if (ttVeto.length>0) {{
        ttProc.push('Veto→0');
        ttScores[tt]=0;
        ttResults.push({{'所属业务层':layer,'模块':mod,'测试类型':tt,'测试类型得分':0,'计算过程':ttProc.join('\\n')}});
        continue;
      }}
      let tw = ttT.reduce((s,t)=>s+getTaskWeight(t.taskType),0);
      ttT.forEach(t => t.nw = tw>0 ? getTaskWeight(t.taskType)/tw : 1/ttT.length);
      let ws = ttT.reduce((s,t)=>s+t.finalScore*t.nw,0);
      ttProc.push(ttT.map(t=>t.taskName+'='+t.finalScore+'×'+t.nw.toFixed(4)).join(', '));
      ttProc.push('加权='+ws.toFixed(4));
      const ttSV = ttT.some(t=>t.flag==='SoftVeto');
      if (ttSV&&ws>=3.75) {{ ws*=0.8; ttProc.push('SoftVeto×0.8='+ws.toFixed(4)); }}
      const ttDC = ttT.some(t=>t.flag==='D_Cap'), ttCC = ttT.some(t=>t.flag==='C_Cap');
      if (ttDC) {{ ws=Math.min(ws,1.9); ttProc.push('Cap(D)→min(,1.9)='+ws.toFixed(2)); }}
      else if (ttCC) {{ ws=Math.min(ws,2.9); ttProc.push('Cap(C)→min(,2.9)='+ws.toFixed(2)); }}
      ttScores[tt]=ws;
      ttProc.push('最终='+ws.toFixed(4));
      ttResults.push({{'所属业务层':layer,'模块':mod,'测试类型':tt,'测试类型得分':Math.round(ws*10000)/10000,'计算过程':ttProc.join('\\n')}});
      procLines.push(tt+'='+ws.toFixed(4));
    }}

    // 模块加权
    const availTT = Object.keys(ttScores);
    if (availTT.length===0) {{ modResults.push({{'所属业务层':layer,'模块':mod,'模块最终评分':0,'模块最终评级':'D','计算过程':'无测试类型得分'}}); continue; }}
    let ttw = availTT.reduce((s,tt)=>s+getTestTypeWeight(tt),0);
    const ttNW = {{}};
    availTT.forEach(tt => ttNW[tt] = ttw>0 ? getTestTypeWeight(tt)/ttw : 1/availTT.length);
    let mws = availTT.reduce((s,tt)=>s+ttScores[tt]*ttNW[tt],0);
    procLines.push('【模块加权】'+availTT.map(tt=>ttScores[tt].toFixed(4)+'×'+ttNW[tt].toFixed(4)).join(' + ')+' = '+mws.toFixed(4));
    let svApplied = false;
    if (hasSV&&mws>=3.75) {{ mws*=0.8; procLines.push('SoftVeto×0.8='+mws.toFixed(4)); svApplied=true; }}
    let mfs = capLim!==null ? Math.min(mws,capLim) : mws;
    procLines.push((capLim?'Cap→min(,'+capLim+')=':'')+'最终='+mfs.toFixed(2));
    const mGrade = getRating(mfs, capType, svApplied);
    procLines.push('评级：'+mGrade);
    modResults.push({{'所属业务层':layer,'模块':mod,'模块最终评分':Math.round(mfs*100)/100,'模块最终评级':mGrade,'计算过程':procLines.join('\\n')}});
  }}

  // 层级
  const layGroups = {{}};
  modResults.forEach(r => {{ if(!layGroups[r['所属业务层']])layGroups[r['所属业务层']]=[]; layGroups[r['所属业务层']].push(r); }});
  const layResults = [];
  for (const lay in layGroups) {{
    const lms = layGroups[lay]; const proc = [];
    const veto = lms.filter(m=>String(m['模块最终评级']).includes('Veto'));
    if (veto.length>0) {{ proc.push('Veto模块('+veto.map(m=>m['模块']).join(',')+')'); layResults.push({{'所属业务层':lay,'层级最终评分':0,'层级最终评级':'D（Veto）','计算过程':proc.join('\\n')}}); continue; }}
    const capD=lms.filter(m=>String(m['模块最终评级']).includes('D（Cap')), capC=lms.filter(m=>String(m['模块最终评级']).includes('C（Cap'));
    let cl=null,ct=null;
    if (capD.length>0) {{ cl=1.9;ct='D_Cap';proc.push('D(Cap)层级→1.9'); }}
    else if (capC.length>0) {{ cl=2.9;ct='C_Cap';proc.push('C(Cap)层级→2.9'); }}
    const mods = lms.map(m=>m['模块']);
    let mw = mods.reduce((s,m)=>s+getModuleWeight(m),0);
    const mnw = {{}};
    mods.forEach(m => mnw[m]=mw>0?getModuleWeight(m)/mw:1/mods.length);
    let lws = lms.reduce((s,r)=>s+r['模块最终评分']*(mnw[r['模块']]||0),0);
    proc.push(lms.map(r=>r['模块最终评分']+'×'+mnw[r['模块']].toFixed(4)).join(' + ')+' = '+lws.toFixed(4));
    let ls = cl!==null ? Math.min(lws,cl) : lws;
    const lr = getRating(ls,ct);
    proc.push('评级：'+lr);
    layResults.push({{'所属业务层':lay,'层级最终评分':Math.round(ls*100)/100,'层级最终评级':lr,'计算过程':proc.join('\\n')}});
  }}

  // 综合
  const totProc = [];
  const vetoLay = layResults.filter(r=>String(r['层级最终评级']).includes('Veto'));
  let totalScore, totalRating;
  if (vetoLay.length>0) {{
    totProc.push('Veto层级('+vetoLay.map(r=>r['所属业务层']).join(',')+')→0');
    totalScore=0; totalRating='D（Veto）';
  }} else {{
    const capDL=layResults.filter(r=>String(r['层级最终评级']).includes('D（Cap')), capCL=layResults.filter(r=>String(r['层级最终评级']).includes('C（Cap'));
    let cl=null,ct=null;
    if (capDL.length>0) {{ cl=1.9;ct='D_Cap';totProc.push('D(Cap)→1.9'); }}
    else if (capCL.length>0) {{ cl=2.9;ct='C_Cap';totProc.push('C(Cap)→2.9'); }}
    const lays=layResults.map(r=>r['所属业务层']);
    let lw=lays.reduce((s,l)=>s+getLayerWeight(l),0);
    const lnw={{}};
    lays.forEach(l=>lnw[l]=lw>0?getLayerWeight(l)/lw:1/lays.length);
    let tws=layResults.reduce((s,r)=>s+r['层级最终评分']*(lnw[r['所属业务层']]||0),0);
    totProc.push(lays.map(l=>(layResults.find(r=>r['所属业务层']===l)['层级最终评分'])+'×'+lnw[l].toFixed(4)).join(' + ')+' = '+tws.toFixed(4));
    totalScore = cl!==null ? Math.min(tws,cl) : tws;
    totalRating = getRating(totalScore,ct);
    totProc.push('评级：'+totalRating);
  }}
  totalScore = Math.round(totalScore*100)/100;

  // 渲染摘要
  document.getElementById('scoreCards').innerHTML =
    '<div class="score-card" style="background:#1a73e8"><h4>综合得分</h4><div class="val">'+totalScore+'</div><div class="lbl">分</div></div>'+
    '<div class="score-card" style="background:#00b42a"><h4>综合评级</h4><div class="val">'+totalRating+'</div><div class="lbl">等级</div></div>'+
    '<div class="score-card" style="background:#17a2b8"><h4>任务总数</h4><div class="val">'+merged.length+'</div><div class="lbl">个</div></div>'+
    '<div class="score-card" style="background:#ff7d00"><h4>已填结果</h4><div class="val">'+merged.filter(t=>t.result).length+'</div><div class="lbl">个</div></div>';

  // 渲染各表
  const ttCols=['所属业务层','模块','测试类型','测试类型得分','计算过程'];
  document.getElementById('stbl-testtype').innerHTML = mkScoreTbl(ttResults, ttCols, null);
  const modCols=['所属业务层','模块','模块最终评分','模块最终评级','计算过程'];
  document.getElementById('stbl-module').innerHTML = mkScoreTbl(modResults, modCols, '模块最终评级');
  const layCols=['所属业务层','层级最终评分','层级最终评级','计算过程'];
  document.getElementById('stbl-layer').innerHTML = mkScoreTbl(layResults, layCols, '层级最终评级');
  document.getElementById('stbl-total').innerHTML = mkScoreTbl([{{'综合最终评分':totalScore,'综合最终评级':totalRating,'计算过程':totProc.join('\\n')}}],['综合最终评分','综合最终评级','计算过程'],'综合最终评级');

  // 统计
  const statsData = [];
  const layers = [...new Set(merged.map(t=>t.layer))];
  layers.forEach(lay => {{
    const ldf = merged.filter(t=>t.layer===lay);
    const mods = [...new Set(ldf.map(t=>t.module))];
    mods.forEach(mod => {{
      const mdf = ldf.filter(t=>t.module===mod);
      const filled = mdf.filter(t=>t.result).length;
      statsData.push({{'所属业务层':lay,'模块':mod,'总数':mdf.length,'已填结果':filled,'未填':mdf.length-filled}});
    }});
  }});
  document.getElementById('stbl-stats').innerHTML = mkScoreTbl(statsData, ['所属业务层','模块','总数','已填结果','未填'], null);

  // 缺失
  const missing = merged.filter(t=>!t.result).map(t=>({{'所属业务层':t.layer,'模块':t.module,'任务ID':t.taskId,'任务名称':t.taskName}}));
  document.getElementById('stbl-missing').innerHTML = missing.length>0 ? mkScoreTbl(missing,['所属业务层','模块','任务ID','任务名称'],null) : '<p style="padding:20px;text-align:center;color:#00b42a">所有任务已填写测试结果</p>';
}}



// ========== 测试报告获取 ==========
let reportCache = null;

function loadLocalReport(input) {{
  const file = input.files[0];
  if (!file) return;
  document.getElementById('reportUrl').value = file.name;
  const reader = new FileReader();
  reader.onload = function(e) {{ parseReportHtml(e.target.result); }};
  reader.readAsText(file, 'utf-8');
}}

function fetchReport() {{
  const url = document.getElementById('reportUrl').value.trim();
  if (!url) {{ showReportMsg('请输入报告地址或选择本地文件', true); return; }}
  showReportMsg('正在获取报告...');
  const proxies = [
    url,
    'https://corsproxy.io/?' + encodeURIComponent(url),
    'https://api.allorigins.win/raw?url=' + encodeURIComponent(url),
    'https://api.codetabs.com/v1/proxy?quest=' + encodeURIComponent(url),
  ];
  let chain = Promise.reject();
  for (const p of proxies) {{
    chain = chain.catch(() => fetch(p).then(r => {{
      if (!r.ok) throw new Error('HTTP ' + r.status);
      return r.text();
    }}));
  }}
  chain.then(html => parseReportHtml(html))
    .catch(e => showReportMsg('URL获取失败（跨域限制），请点击"选择本地文件"按钮上传报告HTML', true));
}}

function parseReportHtml(html) {{
    const parser = new DOMParser();
    const doc = parser.parseFromString(html, 'text/html');

    // 提取被测对象
    const dutSel = doc.getElementById('dutName');
    const duts = dutSel ? [...dutSel.options].map(o => o.textContent.trim()) : [];
    const dutSelect = document.getElementById('reportDut');
    if (duts.length > 1) {{
      dutSelect.style.display = '';
      dutSelect.innerHTML = duts.map(d => '<option value="'+esc(d)+'">'+esc(d)+'</option>').join('');
    }} else if (duts.length === 1) {{
      dutSelect.style.display = '';
      dutSelect.innerHTML = '<option value="'+esc(duts[0])+'">'+esc(duts[0])+'</option>';
    }} else {{
      dutSelect.style.display = 'none';
    }}

    // 从 thead 中定位"召回率"列：找最后一个"序号"th作为数据起点，再找其后的"召回率"th
    const allTh = doc.querySelectorAll('th');
    let detailStart = 0, recallColIdx = -1;
    allTh.forEach((th, i) => {{
      const txt = th.textContent.trim().split('\\n')[0].trim();
      if (txt.startsWith('序号')) detailStart = i;
      if (txt.startsWith('召回率') && i >= detailStart) recallColIdx = i - detailStart;
    }});

    // 提取场景汇总行 (env_results)
    const rows = doc.querySelectorAll('tr.env_results');
    const data = [];
    if (recallColIdx < 0) {{
      showReportMsg('无法在报告表头中找到"召回率"列', true); return;
    }}
    rows.forEach(row => {{
      const cells = row.querySelectorAll('td');
      if (cells.length <= recallColIdx) return;
      const dut = cells[1] ? cells[1].textContent.trim() : '';
      const scene = cells[2] ? cells[2].textContent.trim() : '';
      const recall = cells[recallColIdx] ? cells[recallColIdx].textContent.trim() : '';
      if (recall) data.push({{ dut, scene, tests:'', recalls:'', recall }});
    }});
    reportCache = data;
    showReportMsg('已解析 ' + data.length + ' 条场景数据，' + duts.length + ' 个被测对象');
    matchAndFill();
}}

function matchAndFill() {{
  if (!reportCache) return;
  const dut = document.getElementById('reportDut').value;
  const filtered = getFiltered();
  const dutData = reportCache.filter(d => !dut || d.dut === dut);
  let matched = 0, unmatched = 0;

  dutData.forEach(rd => {{
    const sceneName = rd.scene;
    // 从场景名提取标识：Scene1-1 -> scene1-1, Scene07 -> scene07, task003 -> task003
    const sceneMatch = sceneName.match(/Scene(\d+(?:-\d+)?)/i);
    const taskMatch = sceneName.match(/^(task\d+)/i);
    if (!sceneMatch && !taskMatch) {{ unmatched++; return; }}
    const sceneId = sceneMatch ? 'scene' + sceneMatch[1] : taskMatch[1].toLowerCase();
    const sceneLower = sceneName.toLowerCase();

    // 匹配策略：找 taskId 中包含对应标识的任务
    let candidates = filtered.filter(t => {{
      const tid = (t.taskId || '').toLowerCase();
      const tidBase = tid.replace(/^(wkp_|localasr_|doa_|cloudasr_)/, '');
      return tidBase.includes(sceneId);
    }});

    // 如果匹配到多个，用关键词进一步区分
    if (candidates.length > 1) {{
      // 区分 单扫/扫拖/待机/噪声
      if (sceneLower.includes('扫拖')) {{
        const c = candidates.find(t => (t.taskName||'').includes('扫拖'));
        if (c) candidates = [c];
      }} else if (sceneLower.includes('单扫')) {{
        const c = candidates.find(t => (t.taskName||'').includes('单扫'));
        if (c) candidates = [c];
      }} else if (sceneLower.includes('噪声') && !sceneLower.includes('扫')) {{
        const c = candidates.find(t => (t.taskName||'').includes('噪声') && !(t.taskName||'').includes('扫'));
        if (c) candidates = [c];
      }} else if (sceneLower.includes('待机') || sceneLower.includes('安静') || sceneLower.includes('quiet')) {{
        const c = candidates.find(t => (t.taskName||'').includes('安静') || (t.taskName||'').includes('待机') && !(t.taskName||'').includes('噪声'));
        if (c) candidates = [c];
      }}
    }}

    if (candidates.length > 0) {{
      const task = candidates[0];
      const idx = tasks.indexOf(task);
      if (idx >= 0) {{
        tasks[idx].actualResult = rd.recall;
        matched++;
      }}
    }} else {{
      unmatched++;
    }}
  }});

  // 功能测试类型未匹配到报告数据的，默认填入 "1"
  let funcDefaulted = 0;
  filtered.forEach(t => {{
    if (t.testType === '功能测试' && !t.actualResult) {{
      const idx = tasks.indexOf(t);
      if (idx >= 0) {{ tasks[idx].actualResult = '1'; funcDefaulted++; }}
    }}
  }});

  render();
  let msg = '已填充 ' + matched + ' 条';
  if (funcDefaulted > 0) msg += '，功能测试默认填入 ' + funcDefaulted + ' 条';
  if (unmatched > 0) msg += '，' + unmatched + ' 条未匹配';
  showReportMsg(msg);
}}

function showReportMsg(msg, isErr) {{
  const el = document.getElementById('reportMsg');
  el.textContent = msg;
  el.className = 'report-msg ' + (isErr ? 'report-err' : 'report-ok');
}}


initFilters(); render(); renderWeights();
</script>
</body>
</html>'''


def main():
    if len(sys.argv) < 2:
        print("用法: python generate_html.py <xlsx文件路径>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"文件不存在: {xlsx_path}")
        sys.exit(1)

    project_name = extract_project_name(xlsx_path)
    data = read_xlsx(xlsx_path)
    html = generate_html(data, project_name)

    html_path = os.path.splitext(xlsx_path)[0] + '.html'
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"生成完成: {html_path} ({len(data['tasks'])} 条任务)")


if __name__ == '__main__':
    main()
