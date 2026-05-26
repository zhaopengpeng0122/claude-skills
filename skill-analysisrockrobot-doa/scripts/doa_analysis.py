import re
import os
import math
import argparse
import json
import glob

GEAR_NAMES = {0: "待机", 1: "单拖", 2: "安静档", 3: "标准档", 4: "强劲档", 5: "MAX", 6: "MAX+"}


def angular_error(detected, expected):
    diff = abs(detected - expected)
    return min(diff, 360 - diff)


def parse_excel(excel_path):
    """Parse Excel ground-truth file.
    Columns: ID, wakeup(是/否), expected_angle, doa_value(数字/无结果), result(成功/失败)
    Returns list of dicts with: id, wakeup, expected, doa_excel, result_excel
    """
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]
    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        test_id, wakeup, expected_angle, doa_val, result = row[:5]
        records.append({
            "id": test_id,
            "wakeup": wakeup,
            "expected": expected_angle,
            "doa_excel": doa_val if doa_val != "无结果" else None,
            "result_excel": result,
        })
    return records


def build_excel_report(excel_records, log_records, output_path, correction, test_info, expected_gear=None):
    """Build report using Excel ground-truth for angle mapping."""
    excel_success = [r for r in excel_records if r["wakeup"] == "是" and r["doa_excel"] is not None]
    n = min(len(excel_success), len(log_records))

    from collections import defaultdict
    angle_groups = defaultdict(list)
    errors_all = []
    seq = defaultdict(int)
    lines = []

    lines.append("# DOA 准确率分析报告（Excel 角度映射修正版）")
    lines.append("")
    lines.append("## 测试条件")
    for k, v in test_info.items():
        lines.append(f"- {k}：{v}")
    lines.append(f"- 角度映射：基于 Excel 记录表（修正唤醒失败导致的错位问题）")
    lines.append(f"- 角度修正：日志 DOA 值 +{correction}°")
    lines.append("")

    total_tests = len(excel_records)
    wake_ok = len([r for r in excel_records if r["wakeup"] == "是"])
    wake_fail = total_tests - wake_ok
    lines.append("## 测试概况")
    lines.append("")
    lines.append("| 指标 | 值 |")
    lines.append("|------|-----|")
    lines.append(f"| 总测试数 | {total_tests} |")
    lines.append(f"| 唤醒成功 | {wake_ok} ({wake_ok/total_tests*100:.1f}%) |")
    lines.append(f"| 唤醒失败 | {wake_fail} ({wake_fail/total_tests*100:.1f}%) |")
    lines.append(f"| DOA 分析数 | {n} |")
    lines.append("")

    # Detail table
    has_gear = any(log_records[i].get("gear") is not None for i in range(n))
    lines.append("## 每角度 DOA 检测明细（含 recordId / inputId）")
    lines.append("")
    if has_gear:
        lines.append("| # | 预期角度 | 检测角度 | 误差 | 档位 | recordId | inputId |")
        lines.append("|---|----------|----------|------|------|----------|---------|")
    else:
        lines.append("| # | 预期角度 | 检测角度 | 误差 | recordId | inputId |")
        lines.append("|---|----------|----------|------|----------|---------|")

    for i in range(n):
        exc = excel_success[i]
        log = log_records[i]
        expected = exc["expected"]
        detected = exc["doa_excel"]
        error = angular_error(detected, expected)
        seq[expected] += 1
        errors_all.append(error)
        angle_groups[expected].append({
            "detected": detected, "error": error,
            "gear": log.get("gear"), "recordId": log["recordId"], "inputId": log["inputId"],
        })
        if has_gear:
            g = log.get("gear")
            gear_str = f"{g}({GEAR_NAMES.get(g, '?')})" if g is not None else "-"
            lines.append(f"| {seq[expected]} | {expected}° | {detected}° | {error}° | {gear_str} | {log['recordId']} | {log['inputId']} |")
        else:
            lines.append(f"| {seq[expected]} | {expected}° | {detected}° | {error}° | {log['recordId']} | {log['inputId']} |")
    lines.append("")

    # Per-angle stats
    lines.append("## 每角度统计")
    lines.append("")
    lines.append("| 预期角度 | 有效条数 | 唤醒率 | 平均误差 | ±15°命中 | ±20°命中 | ±30°命中 |")
    lines.append("|----------|----------|--------|----------|----------|----------|----------|")

    best_angle = worst_angle = None
    best_avg = 999
    worst_avg = -1

    for angle in sorted(angle_groups.keys()):
        group = angle_groups[angle]
        g_errors = [r["error"] for r in group]
        g_avg = sum(g_errors) / len(g_errors)
        g_15 = sum(1 for e in g_errors if e <= 15)
        g_20 = sum(1 for e in g_errors if e <= 20)
        g_30 = sum(1 for e in g_errors if e <= 30)
        total_at = len([r for r in excel_records if r["expected"] == angle])
        success_at = len([r for r in excel_records if r["expected"] == angle and r["wakeup"] == "是"])
        if g_avg < best_avg:
            best_avg = g_avg; best_angle = angle
        if g_avg > worst_avg:
            worst_avg = g_avg; worst_angle = angle
        lines.append(f"| {angle}° | {len(group)} | {success_at}/{total_at} | {g_avg:.1f}° | {g_15}/{len(group)} | {g_20}/{len(group)} | {g_30}/{len(group)} |")
    lines.append("")

    # Overall accuracy
    avg_error = sum(errors_all) / len(errors_all)
    std_error = math.sqrt(sum((e - avg_error)**2 for e in errors_all) / len(errors_all))
    max_error = max(errors_all)
    hit_15 = sum(1 for e in errors_all if e <= 15)
    hit_20 = sum(1 for e in errors_all if e <= 20)
    hit_30 = sum(1 for e in errors_all if e <= 30)
    hit_45 = sum(1 for e in errors_all if e <= 45)

    lines.append("## 整体准确率")
    lines.append("")
    lines.append(f"- 平均误差：**{avg_error:.1f}°**")
    lines.append(f"- 误差标准差：**{std_error:.1f}°**")
    lines.append(f"- 最大误差：**{max_error}°**")
    lines.append(f"- 最佳角度：**{best_angle}°**（平均误差 {best_avg:.1f}°）")
    lines.append(f"- 最差角度：**{worst_angle}°**（平均误差 {worst_avg:.1f}°）")
    lines.append("")
    lines.append("| 容差 | 命中数 | 总数 | 准确率 |")
    lines.append("|------|--------|------|--------|")
    lines.append(f"| ±15° | {hit_15} | {n} | {hit_15/n*100:.1f}% |")
    lines.append(f"| ±20° | {hit_20} | {n} | {hit_20/n*100:.1f}% |")
    lines.append(f"| ±30° | {hit_30} | {n} | {hit_30/n*100:.1f}% |")
    lines.append(f"| ±45° | {hit_45} | {n} | {hit_45/n*100:.1f}% |")
    lines.append("")

    # Error distribution
    lines.append("## 误差分布")
    lines.append("")
    buckets = [(0, 5), (6, 10), (11, 15), (16, 30), (31, 45), (46, 90), (91, 180)]
    lines.append("| 误差范围 | 次数 | 占比 |")
    lines.append("|----------|------|------|")
    for lo, hi in buckets:
        count = sum(1 for e in errors_all if lo <= e <= hi)
        lines.append(f"| {lo}°~{hi}° | {count} | {count/n*100:.1f}% |")
    lines.append("")

    # Gear analysis
    if has_gear:
        gear_counts = {}
        for i in range(n):
            g = log_records[i].get("gear")
            if g is not None:
                gear_counts[g] = gear_counts.get(g, 0) + 1
        total_with_gear = sum(gear_counts.values())

        lines.append("## 档位分析")
        lines.append("")
        lines.append("### 档位分布")
        lines.append("")
        lines.append("| 档位值 | 档位名称 | 记录数 | 占比 |")
        lines.append("|--------|----------|--------|------|")
        for g in sorted(gear_counts.keys()):
            name = GEAR_NAMES.get(g, "未知")
            cnt = gear_counts[g]
            lines.append(f"| {g} | {name} | {cnt} | {cnt/total_with_gear*100:.1f}% |")
        lines.append("")

        if expected_gear is not None:
            expected_name = GEAR_NAMES.get(expected_gear, str(expected_gear))
            lines.append("### 档位匹配检查")
            lines.append("")
            lines.append(f"- 预期档位：**{expected_gear}（{expected_name}）**")
            match_count = sum(1 for i in range(n) if log_records[i].get("gear") == expected_gear)
            mismatch_count = sum(1 for i in range(n) if log_records[i].get("gear") is not None and log_records[i].get("gear") != expected_gear)
            lines.append(f"- 匹配记录数：{match_count}/{total_with_gear}（{match_count/total_with_gear*100:.1f}%）")
            lines.append(f"- 不匹配记录数：{mismatch_count}/{total_with_gear}（{mismatch_count/total_with_gear*100:.1f}%）")
            lines.append("")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    result = {
        "mode": "excel",
        "total_tests": total_tests,
        "wakeup_success": wake_ok,
        "wakeup_fail": wake_fail,
        "analyzed": n,
        "correction": correction,
        "avg_error": round(avg_error, 1),
        "max_error": max_error,
        "hit_15": hit_15,
        "hit_20": hit_20,
        "hit_30": hit_30,
        "hit_45": hit_45,
        "total": n,
        "best_angle": best_angle,
        "worst_angle": worst_angle,
    }
    if has_gear:
        result["gear_stats"] = {
            "distribution": {str(k): v for k, v in sorted(gear_counts.items())},
            "expected_gear": expected_gear,
            "match_count": match_count if expected_gear is not None else None,
            "mismatch_count": mismatch_count if expected_gear is not None else None,
        }
    return result


def extract_records(log_path):
    pat_doa = re.compile(r'"accurate_doa":\s*(\d+)')
    pat_rec = re.compile(r'"recordId":\s*"([^"]+)"')
    pat_inp = re.compile(r'"inputId":\s*"([^"]+)"')
    pat_gear = re.compile(r'"gear":\s*(\d+)')
    records = []
    with open(log_path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            if "Recv AI speech event:102" not in line:
                continue
            m_doa = pat_doa.search(line)
            m_rec = pat_rec.search(line)
            m_inp = pat_inp.search(line)
            m_gear = pat_gear.search(line)
            if m_doa and m_rec and m_inp:
                records.append({
                    "doa": int(m_doa.group(1)),
                    "recordId": m_rec.group(1),
                    "inputId": m_inp.group(1),
                    "gear": int(m_gear.group(1)) if m_gear else None,
                })
    return records


def build_report(records, output_path, angles, sentences_per_angle,
                 skip, correction, test_info, expected_gear=None):
    used = records[skip:]
    expected_total = len(angles) * sentences_per_angle
    used = used[:expected_total]

    corrected = [(r["doa"] + correction) % 360 for r in used]
    record_ids = [r["recordId"] for r in used]
    input_ids = [r["inputId"] for r in used]
    gears = [r["gear"] for r in used]
    has_gear = any(g is not None for g in gears)

    lines = []
    lines.append("# DOA 准确率分析报告")
    lines.append("")
    lines.append("## 测试条件")
    for k, v in test_info.items():
        lines.append(f"- {k}：{v}")
    lines.append(f"- 角度范围：{angles[0]}° ~ {angles[-1]}°，间隔 {angles[1]-angles[0]}°，每角度 {sentences_per_angle} 句")
    lines.append(f"- 日志 event:102 读数：{len(records)} 条（剔除前 {skip} 条，+{correction}° 修正后分析 {len(corrected)} 条）")
    lines.append("")

    # Per-angle detail table
    lines.append("## 每角度 DOA 检测明细（含 recordId / inputId）")
    lines.append("")
    if has_gear:
        lines.append("| # | 预期角度 | 实际检测角度 | 档位 | recordId | inputId |")
        lines.append("|---|----------|-------------|------|----------|---------|")
    else:
        lines.append("| # | 预期角度 | 实际检测角度 | recordId | inputId |")
        lines.append("|---|----------|-------------|----------|---------|")
    for i, angle in enumerate(angles):
        start = i * sentences_per_angle
        for j in range(sentences_per_angle):
            idx = start + j
            if idx < len(corrected):
                if has_gear:
                    g = gears[idx]
                    gear_str = f"{g}({GEAR_NAMES.get(g, '?')})" if g is not None else "-"
                    lines.append(f"| {j+1} | {angle}° | {corrected[idx]}° | {gear_str} | {record_ids[idx]} | {input_ids[idx]} |")
                else:
                    lines.append(f"| {j+1} | {angle}° | {corrected[idx]}° | {record_ids[idx]} | {input_ids[idx]} |")
            else:
                if has_gear:
                    lines.append(f"| {j+1} | {angle}° | 缺失 | - | - | - |")
                else:
                    lines.append(f"| {j+1} | {angle}° | 缺失 | - | - |")
    lines.append("")

    # Per-angle summary
    lines.append("## 每角度统计")
    lines.append("")
    header = "| 预期角度 | 检测值 | 平均误差 | ±15°命中 | ±20°命中 | ±30°命中 |"
    sep =    "|----------|--------|----------|----------|----------|----------|"
    lines.append(header)
    lines.append(sep)

    all_errors = []
    total_hit15 = 0
    total_hit20 = 0
    total_hit30 = 0
    total_hit45 = 0

    for i, angle in enumerate(angles):
        start = i * sentences_per_angle
        end = min(start + sentences_per_angle, len(corrected))
        detections = corrected[start:end]

        if not detections:
            lines.append(f"| {angle}° | 无数据 | - | 0/0 | 0/0 | 0/0 |")
            continue

        actual_count = len(detections)

        errors = [angular_error(d, angle) for d in detections]
        all_errors.extend(errors)

        avg_err = sum(errors) / len(errors) if errors else 0
        hit15 = sum(1 for e in errors if e <= 15)
        hit20 = sum(1 for e in errors if e <= 20)
        hit30 = sum(1 for e in errors if e <= 30)
        hit45 = sum(1 for e in errors if e <= 45)
        total_hit15 += hit15
        total_hit20 += hit20
        total_hit30 += hit30
        total_hit45 += hit45

        det_str = ", ".join(str(d) for d in detections)
        lines.append(
            f"| {angle}° | {det_str} | {avg_err:.1f}° | "
            f"{hit15}/{actual_count} | {hit20}/{actual_count} | {hit30}/{actual_count} |"
        )

    lines.append("")

    # Overall accuracy
    lines.append("## 整体准确率")
    lines.append("")
    n = len(all_errors)
    avg_all = sum(all_errors) / n if n else 0
    std_all = math.sqrt(sum((e - avg_all) ** 2 for e in all_errors) / n) if n else 0
    max_err = max(all_errors) if all_errors else 0

    lines.append(f"- 平均误差：**{avg_all:.1f}°**")
    lines.append(f"- 误差标准差：**{std_all:.1f}°**")
    lines.append(f"- 最大误差：**{max_err}°**")
    lines.append("")
    lines.append("| 容差 | 命中数 | 总数 | 准确率 |")
    lines.append("|------|--------|------|--------|")
    lines.append(f"| ±15° | {total_hit15} | {n} | {total_hit15/n*100:.1f}% |")
    lines.append(f"| ±20° | {total_hit20} | {n} | {total_hit20/n*100:.1f}% |")
    lines.append(f"| ±30° | {total_hit30} | {n} | {total_hit30/n*100:.1f}% |")
    lines.append(f"| ±45° | {total_hit45} | {n} | {total_hit45/n*100:.1f}% |")
    lines.append("")

    # Error distribution
    lines.append("## 误差分布")
    lines.append("")
    buckets = [(0, 5), (6, 10), (11, 15), (16, 30), (31, 45), (46, 90), (91, 180)]
    lines.append("| 误差范围 | 次数 | 占比 |")
    lines.append("|----------|------|------|")
    for lo, hi in buckets:
        count = sum(1 for e in all_errors if lo <= e <= hi)
        lines.append(f"| {lo}°~{hi}° | {count} | {count/n*100:.1f}% |")
    lines.append("")

    # Gear analysis section
    if has_gear:
        lines.append("## 档位分析")
        lines.append("")

        gear_counts = {}
        for g in gears:
            if g is not None:
                gear_counts[g] = gear_counts.get(g, 0) + 1
        total_with_gear = sum(gear_counts.values())

        lines.append("### 档位分布")
        lines.append("")
        lines.append("| 档位值 | 档位名称 | 记录数 | 占比 |")
        lines.append("|--------|----------|--------|------|")
        for g in sorted(gear_counts.keys()):
            name = GEAR_NAMES.get(g, "未知")
            cnt = gear_counts[g]
            lines.append(f"| {g} | {name} | {cnt} | {cnt/total_with_gear*100:.1f}% |")
        lines.append("")

        lines.append("### 各角度档位分布")
        lines.append("")
        lines.append("| 预期角度 | 档位分布 |")
        lines.append("|----------|----------|")
        for i, angle in enumerate(angles):
            start = i * sentences_per_angle
            end = min(start + sentences_per_angle, len(gears))
            angle_gears = [g for g in gears[start:end] if g is not None]
            if angle_gears:
                dist_parts = []
                for g in sorted(set(angle_gears)):
                    cnt = angle_gears.count(g)
                    name = GEAR_NAMES.get(g, str(g))
                    dist_parts.append(f"{name}({cnt})")
                lines.append(f"| {angle}° | {', '.join(dist_parts)} |")
            else:
                lines.append(f"| {angle}° | 无数据 |")
        lines.append("")

        if expected_gear is not None:
            expected_name = GEAR_NAMES.get(expected_gear, str(expected_gear))
            lines.append("### 档位匹配检查")
            lines.append("")
            lines.append(f"- 预期档位：**{expected_gear}（{expected_name}）**")
            lines.append("")

            match_count = sum(1 for g in gears if g == expected_gear)
            mismatch_count = sum(1 for g in gears if g is not None and g != expected_gear)
            no_gear_count = sum(1 for g in gears if g is None)

            lines.append(f"- 匹配记录数：{match_count}/{total_with_gear}（{match_count/total_with_gear*100:.1f}%）")
            lines.append(f"- 不匹配记录数：{mismatch_count}/{total_with_gear}（{mismatch_count/total_with_gear*100:.1f}%）")
            if no_gear_count > 0:
                lines.append(f"- 无档位数据：{no_gear_count}")
            lines.append("")

            if mismatch_count > 0:
                lines.append("#### 档位不匹配明细")
                lines.append("")
                lines.append("| # | 预期角度 | 实际检测角度 | 实际档位 | 预期档位 |")
                lines.append("|---|----------|-------------|----------|----------|")
                mismatch_idx = 0
                for i, angle in enumerate(angles):
                    start = i * sentences_per_angle
                    for j in range(sentences_per_angle):
                        idx = start + j
                        if idx < len(gears) and gears[idx] is not None and gears[idx] != expected_gear:
                            actual_name = GEAR_NAMES.get(gears[idx], str(gears[idx]))
                            lines.append(f"| {mismatch_idx+1} | {angle}° | {corrected[idx]}° | {gears[idx]}({actual_name}) | {expected_gear}({expected_name}) |")
                            mismatch_idx += 1
                lines.append("")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    result = {
        "total_records": len(records),
        "skip": skip,
        "analyzed": len(corrected),
        "correction": correction,
        "avg_error": round(avg_all, 1),
        "max_error": max_err,
        "hit_15": total_hit15,
        "hit_20": total_hit20,
        "hit_30": total_hit30,
        "hit_45": total_hit45,
        "total": n,
    }
    if has_gear:
        result["gear_stats"] = {
            "distribution": {str(k): v for k, v in sorted(gear_counts.items())},
            "expected_gear": expected_gear,
            "match_count": sum(1 for g in gears if g == expected_gear) if expected_gear is not None else None,
            "mismatch_count": sum(1 for g in gears if g is not None and g != expected_gear) if expected_gear is not None else None,
        }
    return result


def find_log_files(directory):
    # Find all log files, prefer .tmp.new over .pl9 per directory
    tmp_new = {}
    pl9 = {}
    for root, dirs, files in os.walk(directory):
        if "SPEECH_normal.log.pl9.tmp.new" in files:
            tmp_new[root] = os.path.join(root, "SPEECH_normal.log.pl9.tmp.new")
        elif "SPEECH_normal.log.pl9" in files:
            pl9[root] = os.path.join(root, "SPEECH_normal.log.pl9")
    # Merge: .tmp.new takes priority
    found = {}
    for root, path in pl9.items():
        found[root] = path
    for root, path in tmp_new.items():
        found[root] = path
    # Sort by log file's parent directory name (e.g. 000052, 000053) for chronological order
    def sort_key(item):
        return os.path.basename(item[0])
    sorted_found = sorted(found.items(), key=sort_key)
    # Deduplicate by file size
    seen_sizes = set()
    result = []
    for root, path in sorted_found:
        sz = os.path.getsize(path)
        if sz not in seen_sizes:
            seen_sizes.add(sz)
            result.append(path)
    return result


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="DOA accuracy analysis")
    parser.add_argument("directory", help="Test result directory")
    parser.add_argument("--skip", type=int, default=2, help="Skip first N records")
    parser.add_argument("--correction", type=int, default=60, help="DOA correction angle (+N)")
    parser.add_argument("--start-angle", type=int, default=0, help="Start angle")
    parser.add_argument("--end-angle", type=int, default=360, help="End angle")
    parser.add_argument("--interval", type=int, default=30, help="Angle interval")
    parser.add_argument("--sentences", type=int, default=20, help="Sentences per angle")
    parser.add_argument("--expected-gear", type=int, default=None,
                        help="Expected gear value (0-6). If set, adds gear match check to report.")
    parser.add_argument("--excel", type=str, default=None,
                        help="Excel file path for ground-truth angle mapping. "
                             "Auto-detected if .xlsx exists in directory.")
    args = parser.parse_args()

    log_paths = find_log_files(args.directory)
    if not log_paths:
        print(f"未找到日志文件，已搜索：SPEECH_normal.log.pl9.tmp.new, SPEECH_normal.log.pl9")
        exit(1)

    dir_name = os.path.basename(os.path.normpath(args.directory))
    output_path = os.path.join(args.directory, f"{dir_name}_doa_analysis_report.md")
    test_info = {"目录": args.directory}

    records = []
    for lp in log_paths:
        records.extend(extract_records(lp))
    print(f"提取到 {len(records)} 条 event:102 记录（来自 {len(log_paths)} 个日志文件）")

    # Excel mode: auto-detect or use specified path
    excel_path = args.excel
    if excel_path is None:
        xlsx_files = glob.glob(os.path.join(args.directory, "*.xlsx"))
        if xlsx_files:
            excel_path = xlsx_files[0]

    if excel_path and os.path.exists(excel_path):
        print(f"检测到 Excel 文件：{excel_path}")
        excel_records = parse_excel(excel_path)
        result = build_excel_report(
            excel_records, records, output_path,
            args.correction, test_info, args.expected_gear,
        )
    else:
        angles = list(range(args.start_angle, args.end_angle + 1, args.interval))
        result = build_report(records, output_path, angles, args.sentences,
                              args.skip, args.correction, test_info, args.expected_gear)

    print(f"报告已生成：{output_path}")
    print(json.dumps(result, ensure_ascii=False, indent=2))
